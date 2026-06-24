from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

out = Path(r'D:\Quant\outputs')
out.mkdir(parents=True, exist_ok=True)
md_path = out / 'bt_equal_weight_backtest_data_requirements.md'
xlsx_path = out / 'bt_equal_weight_backtest_field_dictionary.xlsx'
csv_path = out / 'bt_equal_weight_backtest_required_fields.csv'

required_fields = [
    ['date', '必需', '交易日期', 'date', '2015-01-05', '必须可解析为日期；同一symbol内升序；用于生成调仓日和价格索引'],
    ['symbol', '必需', '证券/ETF代码', 'string', '512880', '必须保留前导0；同一ETF全程一致；不可混用交易所后缀口径'],
    ['close', '必需', '收盘价/复权收盘价', 'float', '1.234', '必须为正数；当前项目使用前复权qfq；vectorbt用它成交和计算收益'],
    ['name', '必需', 'ETF名称', 'string', '证券ETF', '用于输出信号和报告；不参与计算但脚本会聚合读取'],
    ['bucket', '必需', '资产分类', 'category', 'sector/defensive/benchmark', 'sector参与行业主题排名；defensive作防守资产；benchmark作对照基准'],
    ['theme', '必需', '主题/行业标签', 'string', '证券', '用于报告解释和检查集中度；脚本会读取'],
]

recommended_fields = [
    ['open', '推荐', '开盘价', 'float', '1.230', '若以后改成次日开盘成交会用到'],
    ['high', '推荐', '最高价', 'float', '1.250', '用于波动、异常价格检查'],
    ['low', '推荐', '最低价', 'float', '1.210', '用于波动、异常价格检查'],
    ['volume', '推荐', '成交量', 'float/int', '1234567', '用于流动性筛选'],
    ['amount', '推荐', '成交额', 'float', '123456789', '用于流动性筛选；比成交量更直观'],
    ['pct_change', '推荐', '日涨跌幅', 'float', '1.23', '可用于校验close计算出的收益是否异常'],
    ['turnover_pct', '可选', '换手率', 'float', '2.31', '辅助判断流动性'],
    ['source', '推荐', '数据来源', 'string', 'eastmoney_push2his', '追溯数据来源'],
    ['adjust', '推荐', '复权口径', 'string', 'qfq', '必须避免前复权、后复权、不复权混用'],
    ['eastmoney_fqt', '可选', '东方财富复权参数', 'int', '1', '1通常表示前复权'],
]

output_fields = [
    ['target_weights', 'date + 每个symbol一列', '每日目标权重矩阵；vectorbt size=targetpercent'],
    ['decisions', 'signal_date/top_n/selected_symbols/target_weight_each/reason', '每个调仓信号日的选择原因和目标权重'],
    ['orders', 'vectorbt orders.records_readable', '模拟订单明细、成交价格、数量、费用'],
    ['daily_value', 'date/portfolio_value/nav/daily_return', '每日组合净值和收益'],
    ['metrics', 'total_return/annualized_return/max_drawdown等', '策略表现汇总'],
]

quality_checks = [
    ['字段完整', 'date、symbol、close、name、bucket、theme 都存在'],
    ['日期可解析', 'date 能被 pandas parse_dates 正确解析'],
    ['代码口径一致', 'symbol 用字符串读取，不能把 510300 变成 510300.0'],
    ['无重复键', '同一个 symbol + date 不能重复'],
    ['价格有效', 'close 不为空且大于0'],
    ['分类有效', 'bucket 只能是 sector、defensive、benchmark'],
    ['至少有基准', 'benchmark里至少有510300或一个可用基准'],
    ['至少有进攻资产', 'sector 数量足够，否则TopN无意义'],
    ['至少有防守资产', 'defensive 用于行业ETF无正动量时切换'],
    ['复权统一', '所有ETF close 使用同一种复权口径，当前为 qfq'],
]

wb = Workbook()
ws = wb.active
ws.title = 'required_fields'
ws.append(['field','level','meaning','type','example','rule'])
for row in required_fields:
    ws.append(row)
ws2 = wb.create_sheet('recommended_fields')
ws2.append(['field','level','meaning','type','example','rule'])
for row in recommended_fields:
    ws2.append(row)
ws3 = wb.create_sheet('output_fields')
ws3.append(['output','fields','meaning'])
for row in output_fields:
    ws3.append(row)
ws4 = wb.create_sheet('quality_checks')
ws4.append(['check','pass_criterion'])
for row in quality_checks:
    ws4.append(row)

fill = PatternFill('solid', fgColor='D9EAF7')
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        width = max(10, min(48, max(len(str(c.value)) if c.value is not None else 0 for c in col[:50]) + 2))
        ws.column_dimensions[col[0].column_letter].width = width
wb.save(xlsx_path)

with csv_path.open('w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(['field','level','meaning','type','example','rule'])
    writer.writerows(required_fields)

md = '''bt / vectorbt 等权组合回测所需数据和字段

定位：这是 D:/Quant 项目中 ETF Top2/Top3 等权轮动回测的数据准备说明。
当前回测脚本：D:/Quant/scripts/vectorbt_topn_etf_rotation.py
当前主数据：D:/Quant/data/etf_momentum_daily_eastmoney_qfq.csv

一、最小必需输入文件

主文件：D:/Quant/data/etf_momentum_daily_eastmoney_qfq.csv

这个文件是长表格式，一行代表某只ETF在某个交易日的数据。

最小必需字段：

1. date
含义：交易日期。
要求：能解析为日期；同一ETF内按时间升序；用于生成调仓日和价格索引。
示例：2015-01-05

2. symbol
含义：ETF代码。
要求：必须按字符串读取，保留前导0；同一ETF全程一致。
示例：512880

3. close
含义：收盘价，当前项目使用前复权收盘价。
要求：不能为空，必须大于0；vectorbt用它计算成交、收益、净值。
示例：1.234

4. name
含义：ETF名称。
要求：用于输出信号、报告和复盘解释。
示例：证券ETF

5. bucket
含义：资产分类。
可选值：sector、defensive、benchmark。
作用：
sector：行业/主题ETF，参与动量排名。
defensive：防守资产，行业ETF无正动量时切换。
benchmark：基准资产，用于对照，不参与行业主题排名。

6. theme
含义：主题/行业标签。
作用：解释信号和检查行业集中度。
示例：证券、银行、芯片、沪深300、货币债券。

二、推荐保留字段

这些字段不是当前脚本的最小必需项，但强烈建议保留：

open：开盘价，以后如果改成次日开盘成交会用到。
high：最高价，用于异常价格检查。
low：最低价，用于异常价格检查。
volume：成交量，用于流动性筛选。
amount：成交额，用于流动性筛选，更适合ETF。
pct_change：日涨跌幅，用于校验收益异常。
turnover_pct：换手率，辅助流动性判断。
source：数据来源，方便追溯。
adjust：复权口径，当前应为 qfq。
eastmoney_fqt：东方财富复权参数，当前为 1。

三、宽表价格矩阵

当前项目也有宽表：
D:/Quant/data/etf_momentum_close_wide_qfq.csv

格式：
第一列 date
后面每一列是一个ETF代码
单元格是该ETF当天的前复权收盘价

宽表适合：
看价格矩阵
检查缺失值
快速计算收益和动量

但当前 vectorbt_topn_etf_rotation.py 的主入口读的是长表，不是宽表。

四、字段如何进入回测

脚本读取长表后：

1. 用 date + symbol + close 透视成价格矩阵 close
行：date
列：symbol
值：close

2. 用 bucket 分出三类资产
sector_symbols：参与TopN排名的行业主题ETF
defensive_symbols：防守资产
benchmark_symbols：基准资产

3. 用 close / close.shift(42) - 1 计算42日动量

4. 每月末生成信号
按 sector ETF 的动量从高到低排序。
选择正动量Top2或Top3。
入选资产等权。
如果sector无正动量，则选正动量防守资产。
如果防守资产也无正动量，则空仓。

5. 生成 target_weights
这是每日目标权重矩阵。
vectorbt 使用 size_type='targetpercent'，按目标百分比调仓。

五、必须做的数据质量检查

回测前至少检查：

1. date、symbol、close、name、bucket、theme 是否存在。
2. date 是否能正确解析。
3. symbol 是否是字符串，不能变成 510300.0。
4. 同一 symbol + date 是否重复。
5. close 是否为空或小于等于0。
6. bucket 是否只包含 sector、defensive、benchmark。
7. sector 是否有足够资产，否则TopN无意义。
8. defensive 是否至少有一个资产，否则风控切换失效。
9. benchmark 是否至少有510300或一个可用基准。
10. 所有价格是否统一复权口径，当前项目应统一为 qfq。

六、当前项目数据状态

已检查 D:/Quant/data/etf_momentum_daily_eastmoney_qfq.csv：

行数：37957
字段数：18
日期范围：2015-01-05 到 2026-06-18
ETF数量：18
分类数量：sector 14，benchmark 2，defensive 2
必需字段缺失：0
symbol + date 重复：0
close 小于等于0：0

所以当前主数据可以支持现有 Top2/Top3 等权组合回测。

七、输出文件会有哪些

运行回测后会生成：

D:/Quant/outputs/etf_topn_rotation_vectorbt_metrics.csv
汇总指标。

D:/Quant/outputs/etf_top2_equal_weight_rotation_vectorbt_daily_value.csv
Top2每日净值。

D:/Quant/outputs/etf_top2_equal_weight_rotation_vectorbt_target_weights.csv
Top2每日目标权重。

D:/Quant/outputs/etf_top2_equal_weight_rotation_vectorbt_decisions.csv
Top2调仓信号和原因。

D:/Quant/outputs/etf_top2_equal_weight_rotation_vectorbt_orders.csv
Top2模拟订单。

Top3也有同样一组文件。

八、最容易犯的错

1. 混用前复权、不复权价格。
2. 把 benchmark 也放进 sector 排名，导致基准ETF参与策略选择。
3. symbol 被当数字读入，代码格式被破坏。
4. 新ETF上市前价格缺失没有处理，导致早期动量样本不公平。
5. 只看最终收益，不看最大回撤和样本外稳定性。
6. 用未来数据生成当期信号，造成未来函数。

一句话：
等权组合回测最少需要 date、symbol、close、name、bucket、theme。真正影响结果的是 close 的复权口径、bucket 分类是否正确、调仓日是否没有偷看未来。
'''
md_path.write_text(md, encoding='utf-8')
print(md_path)
print(xlsx_path)
print(csv_path)
