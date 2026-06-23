from pathlib import Path
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

out = Path(r'D:\Quant\outputs')
out.mkdir(parents=True, exist_ok=True)
xlsx = out / 'etf_rotation_trade_log_template.xlsx'
csv_path = out / 'etf_rotation_trade_log_template.csv'
md = out / 'etf_rotation_trade_log_guide.md'

trade_log_headers = [
    'trade_id', 'strategy_version', 'trade_stage', 'trade_date', 'signal_date',
    'symbol', 'asset_name', 'asset_type', 'action', 'risk_state',
    'signal_rank', 'momentum_score', 'target_weight_before', 'target_weight_after',
    'planned_price', 'planned_quantity', 'planned_amount',
    'paper_price', 'paper_quantity', 'gross_amount', 'estimated_cost', 'net_amount',
    'cash_before', 'cash_after', 'nav_before', 'nav_after',
    'portfolio_weight_after', 'benchmark_symbol', 'benchmark_return_since_signal',
    'strategy_return_since_signal', 'excess_return_since_signal',
    'rebalance_reason', 'skip_reason', 'manual_override', 'override_reason',
    'data_quality_status', 'execution_status', 'emotion_state', 'review_note'
]

example_trade_row = [
    'PT-20261008-001', 'candidate_v1_top2_a', 'paper_trade', '2026-10-08', '2026-09-30',
    '159915', '创业板ETF', 'ETF', 'BUY', 'risk_on',
    1, 0.138, 0.00, 0.50,
    2.000, 25000, 50000.00,
    2.000, 24987, 49974.00, 24.99, 49998.99,
    100000.00, 50001.01, 100000.00, 99975.01,
    0.50, '510300', '', '', '',
    '月末动量Top2，risk_on，按规则调仓', '', '否', '',
    'ok', 'filled_paper', '平静', '示例行：真实使用时可删除'
]

sheets = {
    'trade_log': [trade_log_headers],
    'example_trade_log': [trade_log_headers, example_trade_row],
    'signal_snapshot': [[
        'signal_date', 'strategy_version', 'risk_state', 'benchmark_symbol',
        'benchmark_close', 'benchmark_ma120', 'risk_rule_passed',
        'symbol', 'asset_name', 'momentum_score', 'rank', 'selected',
        'target_weight', 'signal_note'
    ]],
    'order_log': [[
        'order_id', 'trade_id', 'order_date', 'symbol', 'asset_name', 'action',
        'order_type', 'planned_price', 'planned_quantity', 'planned_amount',
        'paper_fill_price', 'paper_fill_quantity', 'gross_amount', 'estimated_cost',
        'net_amount', 'slippage_assumption', 'order_status', 'note'
    ]],
    'portfolio_snapshot': [[
        'snapshot_date', 'strategy_version', 'symbol', 'asset_name', 'quantity',
        'close_price', 'market_value', 'weight', 'target_weight', 'weight_deviation',
        'cash', 'total_nav', 'risk_state', 'note'
    ]],
    'nav_daily': [[
        'date', 'strategy_version', 'nav', 'daily_return', 'benchmark_symbol',
        'benchmark_close', 'benchmark_daily_return', 'excess_daily_return',
        'drawdown_from_peak', 'cash_weight', 'offensive_weight', 'defensive_weight',
        'note'
    ]],
    'weekly_review': [[
        'week_id', 'week_start', 'week_end', 'strategy_version', 'holding_summary',
        'strategy_week_return', 'benchmark_week_return', 'excess_week_return',
        'max_drawdown_week', 'risk_state_change', 'rule_followed', 'manual_override_count',
        'main_event', 'lesson', 'next_week_action'
    ]],
    'exception_log': [[
        'exception_id', 'date', 'exception_type', 'description', 'affected_symbol',
        'severity', 'action_taken', 'resolved', 'resolution_note'
    ]],
    'checklist': [
        ['stage', 'check_item', 'done'],
        ['交易前', '确认信号日期和策略版本', ''],
        ['交易前', '检查ETF价格、复权、停牌、缺失值', ''],
        ['交易前', '检查510300风控状态', ''],
        ['交易前', '确认目标仓位和当前仓位差异', ''],
        ['交易前', '写明是否人工覆盖信号', ''],
        ['交易中', '按纸面成交规则记录价格和数量', ''],
        ['交易中', '扣除估算成本', ''],
        ['交易后', '更新现金、持仓、NAV', ''],
        ['交易后', '记录是否偏离计划', ''],
        ['周末', '更新收益、超额、回撤和经验', ''],
    ],
    'enum_values': [
        ['trade_stage', 'action', 'risk_state', 'manual_override', 'data_quality_status', 'execution_status', 'emotion_state'],
        ['research_simulation', 'BUY', 'risk_on', '否', 'ok', 'planned', '平静'],
        ['paper_trade', 'SELL', 'defensive', '是', 'missing_price', 'filled_paper', '焦虑'],
        ['live_candidate', 'HOLD', 'cash', '', 'abnormal_price', 'skipped', '贪婪'],
        ['live_trade', 'SKIP', 'unknown', '', 'suspended', 'partial_paper', '犹豫'],
        ['', 'REBALANCE', '', '', 'need_check', 'cancelled', '冲动'],
        ['', '', '', '', '', '', '复盘后确认'],
    ],
    'field_description': [
        ['field', 'description'],
        ['trade_id', '交易唯一编号，建议格式 PT-日期-序号。'],
        ['strategy_version', '策略版本，例如 candidate_v1_top2_a。'],
        ['trade_stage', '交易阶段：研究模拟、纸面交易、实盘候选、实盘。'],
        ['trade_date', '纸面交易执行日期。'],
        ['signal_date', '生成信号的日期，通常是月末收盘后。'],
        ['symbol', 'ETF代码。'],
        ['action', 'BUY/SELL/HOLD/SKIP/REBALANCE。'],
        ['risk_state', '风控状态：risk_on/defensive/cash。'],
        ['signal_rank', '信号排名。'],
        ['momentum_score', '动量分数，来自策略规则。'],
        ['target_weight_after', '交易后目标权重。'],
        ['paper_price', '纸面成交价格，必须写明口径，例如次日收盘价。'],
        ['estimated_cost', '估算交易成本，纸面交易也必须扣除。'],
        ['nav_before/nav_after', '交易前后纸面账户净值。'],
        ['manual_override', '是否人工覆盖策略信号。'],
        ['override_reason', '如果人工覆盖，必须写原因。'],
        ['data_quality_status', '数据是否可靠。'],
        ['execution_status', '纸面执行状态。'],
        ['emotion_state', '记录交易时情绪，防止策略被心态污染。'],
        ['review_note', '盘后复盘备注。'],
    ],
}

wb = Workbook()
ws = wb.active
ws.title = 'trade_log'
for row in sheets.pop('trade_log'):
    ws.append(row)
for title, rows in sheets.items():
    ws = wb.create_sheet(title)
    for row in rows:
        ws.append(row)

header_fill = PatternFill('solid', fgColor='D9EAF7')
for ws in wb.worksheets:
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 10
        for cell in col[:30]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[col_letter].width = max_len + 2

ws = wb['trade_log']
header_to_col = {cell.value: cell.column_letter for cell in ws[1]}
validations = {
    'trade_stage': ['research_simulation', 'paper_trade', 'live_candidate', 'live_trade'],
    'action': ['BUY', 'SELL', 'HOLD', 'SKIP', 'REBALANCE'],
    'risk_state': ['risk_on', 'defensive', 'cash', 'unknown'],
    'manual_override': ['否', '是'],
    'data_quality_status': ['ok', 'missing_price', 'abnormal_price', 'suspended', 'need_check'],
    'execution_status': ['planned', 'filled_paper', 'skipped', 'partial_paper', 'cancelled'],
    'emotion_state': ['平静', '焦虑', '贪婪', '犹豫', '冲动', '复盘后确认'],
}
for field, values in validations.items():
    col = header_to_col.get(field)
    if col:
        dv = DataValidation(type='list', formula1='"' + ','.join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f'{col}2:{col}1000')

wb.save(xlsx)

with csv_path.open('w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f)
    writer.writerow(trade_log_headers)

md.write_text('''ETF动量轮动交易日志模板使用说明

定位：这是研究和纸面交易日志，不是实盘交易建议。

一、每次交易至少填什么

最少要填 trade_log 里的这些字段：

1. trade_id：交易编号。
2. strategy_version：策略版本。
3. trade_stage：阶段，建议先用 paper_trade。
4. trade_date：纸面执行日期。
5. signal_date：信号生成日期。
6. symbol / asset_name：ETF代码和名称。
7. action：BUY、SELL、HOLD、SKIP或REBALANCE。
8. risk_state：risk_on、defensive或cash。
9. target_weight_after：交易后目标权重。
10. paper_price、paper_quantity、estimated_cost：纸面成交记录和成本。
11. nav_before、nav_after：纸面账户净值变化。
12. manual_override：是否人工覆盖信号。
13. review_note：盘后备注。

二、推荐填写顺序

1. signal_snapshot：先记录信号来源和排名。
2. trade_log：记录本次交易主表。
3. order_log：如果有买卖动作，记录模拟订单。
4. portfolio_snapshot：更新交易后持仓。
5. nav_daily：每天或每周更新净值。
6. weekly_review：每周末复盘。
7. exception_log：记录异常、缺失数据、人工覆盖等问题。

三、纸面交易纪律

1. 纸面交易也要扣成本。
2. 不允许事后改信号日期。
3. 不允许因为涨跌随意改规则。
4. 人工覆盖可以记录，但必须写原因。
5. 如果数据缺失或价格异常，默认不交易。

四、关键判断

交易日志不是为了证明自己对了，而是为了保留证据。
连续12周后，用这些日志计算：

- 总收益
- 相对510300超额收益
- 最大回撤
- 信号执行率
- 人工覆盖次数
- 规则修改次数
- 情绪压力

然后再决定：继续研究、纸面交易强化，还是暂缓。

文件：

- D:/Quant/outputs/etf_rotation_trade_log_template.xlsx
- D:/Quant/outputs/etf_rotation_trade_log_template.csv
- D:/Quant/outputs/etf_rotation_trade_log_guide.md
''', encoding='utf-8')

print(xlsx)
print(csv_path)
print(md)
